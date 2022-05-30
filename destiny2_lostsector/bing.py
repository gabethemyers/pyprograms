import time
import smtplib, ssl
from openpyxl import load_workbook
import illegal_stuff

def emailSend(receiver_email, message):
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(illegal_stuff.email_username, illegal_stuff.email_pass)
        server.sendmail(illegal_stuff.email_username, receiver_email, message)

# intitalizing workbook and sheet, also getting todays date.
workbook = load_workbook(filename='Lost Sector.xlsx')
sheet = workbook.active
date = time.strftime('%m/%d/%Y')
index = 1

# iterating through all rows in column 1 and seeing if todays date matches the sheetdate.
for value in sheet.iter_rows(min_row=2, max_row=183, min_col=1, max_col=1, values_only=True):
    index += 1
    # returns a tuple but I only need the first element
    sheetdate = value[0].strftime('%m/%d/%Y')
    if date == sheetdate:
        break

lost_sector_name = sheet.cell(row=index, column=3).value
master_lost_sector_loot = sheet.cell(row=index, column=6).value

if lost_sector_name == 'Chamber of Starlight (DC)':
    emailSend(illegal_stuff.my_email, 'Today\'s lost sector is Chamber of Starlight (DC)')
if lost_sector_name == 'Perdition (Europa)':
    emailSend(illegal_stuff.my_email, 'Today\'s lost sector is Perdition (Europa)')
if master_lost_sector_loot == 'Legs':
    emailSend(illegal_stuff.my_email, 'Today\'s master lost sector loot is legs')
